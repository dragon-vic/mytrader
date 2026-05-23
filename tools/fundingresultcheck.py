from __future__ import annotations

import hashlib
import hmac
import os
import platform
import sys
import time
from datetime import datetime
from datetime import timezone
from decimal import Decimal
from decimal import getcontext
from pathlib import Path
from statistics import mean
from typing import Any
from urllib.parse import urlencode

import requests
import yaml
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font


getcontext().prec = 28
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
BASE_URL = "https://fapi.binance.com"
DAY_MS = 24 * 60 * 60 * 1000
HOUR_MS = 60 * 60 * 1000
REQUEST_DELAY_SEC = 0.12


# 转成本地时间字符串，便于和交易所页面核对。
def format_time(ms: int | None) -> str:
    if ms is None:
        return "-"
    dt = datetime.fromtimestamp(ms / 1000, tz=timezone.utc).astimezone()
    return dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]


def format_hour(ms: int) -> str:
    dt = datetime.fromtimestamp(ms / 1000, tz=timezone.utc).astimezone()
    return dt.strftime("%Y-%m-%d %H:00")


def fmt(value: Decimal | float | int | None, places: int = 4) -> str:
    if value is None:
        return "-"
    if isinstance(value, Decimal):
        return f"{value:.{places}f}"
    return f"{value:.{places}f}" if isinstance(value, float) else str(value)


def offset_label(offset_ms: int | None) -> str:
    if offset_ms is None:
        return "-"
    sign = "+" if offset_ms >= 0 else "-"
    return f"T{sign}{abs(offset_ms)}ms"


# 读取 maxfunding 当前配置，用 exclude_symbols 限定策略结果范围。
def load_maxfunding() -> dict[str, Any]:
    with (ROOT / "config" / "maxfunding.yaml").open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# 构建带实盘 U 本位 API key 的 session。
def build_session() -> requests.Session:
    load_dotenv(ROOT / ".env")
    session = requests.Session()
    session.headers.update({"X-MBX-APIKEY": os.environ["BINANCE_FUTURES_API_KEY"]})
    if platform.system() == "Windows":
        proxy_url = os.environ.get("PROXY_URL")
        if proxy_url:
            session.proxies.update({"http": proxy_url, "https": proxy_url})
    return session


# 发 Binance 签名请求。
def signed_request(
    session: requests.Session,
    secret: bytes,
    path: str,
    params: dict[str, Any] | None = None,
) -> list[dict[str, Any]] | dict[str, Any]:
    payload = dict(params or {})
    payload["timestamp"] = int(time.time() * 1000)
    payload["recvWindow"] = 10000
    query = urlencode(payload)
    sig = hmac.new(secret, query.encode(), hashlib.sha256).hexdigest()
    response = session.get(f"{BASE_URL}{path}?{query}&signature={sig}", timeout=30)
    data = response.json()
    if response.status_code in (418, 429):
        raise RuntimeError(f"RATE_LIMIT HTTP {response.status_code}: {data}")
    if response.status_code >= 400:
        raise RuntimeError(f"HTTP {response.status_code}: {data}")
    return data


# 发 Binance 公共行情请求。
def public_request(
    session: requests.Session,
    path: str,
    params: dict[str, Any] | None = None,
) -> list[dict[str, Any]] | list[list[Any]]:
    response = session.get(f"{BASE_URL}{path}", params=params or {}, timeout=30)
    data = response.json()
    if response.status_code in (418, 429):
        raise RuntimeError(f"RATE_LIMIT HTTP {response.status_code}: {data}")
    if response.status_code >= 400:
        raise RuntimeError(f"HTTP {response.status_code}: {data}")
    return data


# 按时间分页读取账户接口。
def signed_pages(
    session: requests.Session,
    secret: bytes,
    path: str,
    params: dict[str, Any],
    start_ms: int,
    end_ms: int,
    key: str = "time",
    max_pages: int = 100,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current = start_ms
    for _ in range(max_pages):
        payload = dict(params)
        payload.update({"startTime": current, "endTime": end_ms, "limit": 1000})
        batch = signed_request(session, secret, path, payload)
        if not isinstance(batch, list) or not batch:
            break
        batch.sort(key=lambda item: int(item[key]))
        rows.extend(batch)
        last = int(batch[-1][key])
        if len(batch) < 1000 or last >= end_ms:
            break
        current = last + 1
        time.sleep(REQUEST_DELAY_SEC)

    seen: set[tuple[Any, ...]] = set()
    result = []
    for row in rows:
        ident = (
            row.get("symbol"),
            row.get("id"),
            row.get("orderId"),
            row.get("time"),
            row.get("incomeType"),
            row.get("tranId"),
        )
        if ident not in seen:
            seen.add(ident)
            result.append(row)
    return sorted(result, key=lambda item: int(item[key]))


# userTrades 单次最多 7 天，按窗口切开。
def user_trades(
    session: requests.Session,
    secret: bytes,
    symbol: str,
    start_ms: int,
    end_ms: int,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current = start_ms
    span = 7 * DAY_MS - 60_000
    while current <= end_ms:
        part_end = min(end_ms, current + span)
        rows.extend(
            signed_pages(
                session,
                secret,
                "/fapi/v1/userTrades",
                {"symbol": symbol},
                current,
                part_end,
            )
        )
        current = part_end + 1
        time.sleep(REQUEST_DELAY_SEC)

    seen: set[tuple[Any, ...]] = set()
    result = []
    for row in rows:
        ident = (row.get("symbol"), row.get("id"), row.get("orderId"), row.get("time"))
        if ident not in seen:
            seen.add(ident)
            result.append(row)
    return sorted(result, key=lambda item: int(item["time"]))


# 按 Binance 成交方向转成持仓增量。
def trade_delta(trade: dict[str, Any]) -> Decimal:
    qty = Decimal(str(trade["qty"]))
    side = trade["side"]
    position_side = trade.get("positionSide", "BOTH")
    if position_side == "SHORT":
        return qty if side == "SELL" else -qty
    return qty if side == "BUY" else -qty


# 从成交记录重建从非零到归零的持仓区间。
def closed_segments(
    trades: list[dict[str, Any]],
    start_ms: int,
    end_ms: int,
) -> list[dict[str, Any]]:
    segments = []
    sides = sorted({trade.get("positionSide", "BOTH") for trade in trades})
    for position_side in sides:
        pos = Decimal("0")
        current: dict[str, Any] | None = None
        side_trades = [
            trade for trade in trades
            if trade.get("positionSide", "BOTH") == position_side
        ]
        for trade in side_trades:
            before = pos
            after = before + trade_delta(trade)
            if before == 0 and after != 0:
                current = {
                    "symbol": trade["symbol"],
                    "positionSide": position_side,
                    "direction": "LONG" if after > 0 else "SHORT",
                    "openTime": int(trade["time"]),
                    "trades": [],
                }
            if current is not None:
                current["trades"].append(trade)

            flipped = before != 0 and (
                (before > 0 and after < 0) or (before < 0 and after > 0)
            )
            if before != 0 and (after == 0 or flipped) and current is not None:
                current["closeTime"] = int(trade["time"])
                segments.append(current)
                current = (
                    {
                        "symbol": trade["symbol"],
                        "positionSide": position_side,
                        "direction": "LONG" if after > 0 else "SHORT",
                        "openTime": int(trade["time"]),
                        "trades": [trade],
                    }
                    if after != 0
                    else None
                )
            pos = after

    return sorted(
        [
            segment for segment in segments
            if start_ms <= int(segment["closeTime"]) <= end_ms
        ],
        key=lambda item: int(item["closeTime"]),
    )


# 汇总一个持仓区间的仓位、PnL 和手续费。
def summarize_segment(segment: dict[str, Any]) -> dict[str, Any]:
    pos = Decimal("0")
    max_qty = Decimal("0")
    max_notional = Decimal("0")
    pnl = Decimal("0")
    fee = Decimal("0")
    for trade in segment["trades"]:
        pos += trade_delta(trade)
        max_qty = max(max_qty, abs(pos))
        notional = abs(pos) * Decimal(str(trade["price"]))
        max_notional = max(max_notional, notional)
        pnl += Decimal(str(trade.get("realizedPnl", "0")))
        fee += Decimal(str(trade.get("commission", "0")))
    return {
        **segment,
        "maxQty": max_qty,
        "notional": max_notional,
        "realizedPnl": pnl,
        "commission": fee,
    }


# 取开仓后最近的小时整点，作为策略对应的 funding 时间 T。
def funding_time(open_ms: int) -> int:
    return ((open_ms + HOUR_MS - 1) // HOUR_MS) * HOUR_MS


# 从 30 天 funding 入账里找当前 symbol 在 T 附近的记录。
def match_funding(
    funding_rows: list[dict[str, Any]],
    symbol: str,
    target_ms: int,
    tolerance_ms: int,
) -> tuple[bool, Decimal, int | None]:
    candidates = [
        row for row in funding_rows
        if row.get("symbol") == symbol
        and abs(int(row["time"]) - target_ms) <= tolerance_ms
        and Decimal(str(row.get("income", "0"))) != 0
    ]
    if not candidates:
        return False, Decimal("0"), None
    total = sum((Decimal(str(row["income"])) for row in candidates), Decimal("0"))
    first_time = min(int(row["time"]) for row in candidates)
    return True, total, first_time


# 拉 funding 前后的 aggTrades。
def agg_trades(
    session: requests.Session,
    symbol: str,
    start_ms: int,
    end_ms: int,
    max_pages: int = 20,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current = start_ms
    for _ in range(max_pages):
        batch = public_request(
            session,
            "/fapi/v1/aggTrades",
            {
                "symbol": symbol,
                "startTime": current,
                "endTime": end_ms,
                "limit": 1000,
            },
        )
        if not isinstance(batch, list) or not batch:
            break
        rows.extend(batch)
        last = max(int(row["T"]) for row in batch)
        if len(batch) < 1000 or last >= end_ms:
            break
        current = last + 1
        time.sleep(REQUEST_DELAY_SEC)
    return sorted(rows, key=lambda row: int(row["T"]))


# 用 aggTrades 重建逐秒 high/low/close。
def second_bars(
    trades: list[dict[str, Any]],
    start_ms: int,
    end_ms: int,
) -> dict[int, dict[str, float]]:
    bars: dict[int, dict[str, float]] = {}
    for trade in trades:
        trade_ms = int(trade["T"])
        if trade_ms < start_ms or trade_ms > end_ms:
            continue
        second_ms = (trade_ms // 1000) * 1000
        price = float(trade["p"])
        bar = bars.get(second_ms)
        if bar is None:
            bars[second_ms] = {"high": price, "low": price, "close": price}
        else:
            bar["high"] = max(bar["high"], price)
            bar["low"] = min(bar["low"], price)
            bar["close"] = price
    return bars


def second_range_bps(bar: dict[str, float] | None) -> float:
    if bar is None:
        return 0.0
    high = bar["high"]
    low = bar["low"]
    close = bar["close"]
    if close == 0:
        return 0.0
    return (high - low) / close * 10000


def total_range_bps(bars: list[dict[str, float]]) -> float | None:
    if not bars:
        return None
    high = max(bar["high"] for bar in bars)
    low = min(bar["low"] for bar in bars)
    close = bars[-1]["close"]
    if close == 0:
        return None
    return (high - low) / close * 10000


# 计算 funding 前 30 秒和 funding 后 3 秒逐秒波动。
def volatility_stats(
    bars: dict[int, dict[str, float]],
    target_ms: int,
    before_sec: int,
) -> dict[str, float | None]:
    pre_seconds = [
        target_ms - second * 1000
        for second in range(before_sec, 0, -1)
    ]
    pre_bars = [bars[second] for second in pre_seconds if second in bars]
    pre_ranges = [second_range_bps(bars.get(second)) for second in pre_seconds]
    post = {}
    for index in range(3):
        second_ms = target_ms + index * 1000
        post[f"post{index + 1}"] = second_range_bps(bars.get(second_ms))

    return {
        "preAvg": mean(pre_ranges) if pre_ranges else None,
        "preMax": max(pre_ranges) if pre_ranges else None,
        "preTotal": total_range_bps(pre_bars),
        **post,
    }


def allowed_symbols(config: dict[str, Any]) -> set[str]:
    params = config["strategy"]["params"]
    excluded = params.get("exclude_symbols", [])
    return {f"{str(symbol).upper()}USDT" for symbol in excluded}


def amount_bps(amount: Decimal, notional: Decimal) -> float | None:
    if notional == 0:
        return None
    return float(amount / notional * Decimal("10000"))


def write_excel(rows: list[dict[str, Any]], output_path: Path) -> None:
    headers = [
        "序号",
        "T时间",
        "标的",
        "方向",
        "开偏",
        "平偏",
        "持仓ms",
        "拿到funding",
        "名义价值",
        "funding_bps",
        "pnl_bps",
        "手续费_bps",
        "前30秒均波bps",
        "前30秒最大bps",
        "前30秒整体bps",
        "后第1秒bps",
        "后第2秒bps",
        "后第3秒bps",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "funding结果"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for index, row in enumerate(rows, 1):
        notional = row["notional"]
        ws.append([
            index,
            format_hour(row["fundingTime"]),
            row["symbol"],
            row["direction"],
            offset_label(row["openOffsetMs"]),
            offset_label(row["closeOffsetMs"]),
            row["closeTime"] - row["openTime"],
            "是" if row["fundingHit"] else "否",
            float(notional),
            amount_bps(row["fundingAmount"], notional),
            amount_bps(row["realizedPnl"], notional),
            amount_bps(row["commission"], notional),
            row["preAvg"],
            row["preMax"],
            row["preTotal"],
            row["post1"],
            row["post2"],
            row["post3"],
        ])

    for row in ws.iter_rows(min_row=2, min_col=9, max_col=18):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0"

    widths = {
        "A": 6,
        "B": 18,
        "C": 14,
        "D": 8,
        "E": 10,
        "F": 10,
        "G": 10,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 10,
        "L": 12,
        "M": 14,
        "N": 14,
        "O": 14,
        "P": 12,
        "Q": 12,
        "R": 12,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    wb.save(output_path)


def main(
    days: int,
    open_before_hour_ms: int,
    before_sec: int,
    after_sec: int,
    funding_tolerance_ms: int,
    include_excluded: bool,
    output_path: Path,
) -> None:
    config = load_maxfunding()
    excluded = allowed_symbols(config)
    session = build_session()
    secret = os.environ["BINANCE_FUTURES_API_SECRET"].encode()

    now_ms = int(time.time() * 1000)
    start_ms = now_ms - days * DAY_MS
    income_rows = signed_pages(
        session,
        secret,
        "/fapi/v1/income",
        {"incomeType": "REALIZED_PNL"},
        start_ms,
        now_ms,
    )
    funding_rows = signed_pages(
        session,
        secret,
        "/fapi/v1/income",
        {"incomeType": "FUNDING_FEE"},
        start_ms,
        now_ms,
    )

    symbols = sorted({
        row["symbol"] for row in income_rows
        if Decimal(str(row.get("income", "0"))) != 0
    })
    if not include_excluded:
        symbols = [symbol for symbol in symbols if symbol not in excluded]

    results = []
    for symbol in symbols:
        trades = user_trades(session, secret, symbol, start_ms, now_ms)
        for segment in closed_segments(trades, start_ms, now_ms):
            summary = summarize_segment(segment)
            target_ms = funding_time(summary["openTime"])
            open_offset = summary["openTime"] - target_ms
            if not -open_before_hour_ms <= open_offset <= 0:
                continue

            market_trades = agg_trades(
                session,
                symbol,
                target_ms - before_sec * 1000,
                target_ms + after_sec * 1000 - 1,
            )
            bars = second_bars(
                market_trades,
                target_ms - before_sec * 1000,
                target_ms + after_sec * 1000 - 1,
            )
            stats = volatility_stats(bars, target_ms, before_sec)
            if summary["openTime"] <= target_ms <= summary["closeTime"]:
                hit, funding_amount, funding_ms = match_funding(
                    funding_rows,
                    symbol,
                    target_ms,
                    funding_tolerance_ms,
                )
            else:
                hit, funding_amount, funding_ms = False, Decimal("0"), None
            results.append({
                **summary,
                **stats,
                "fundingTime": target_ms,
                "fundingHit": hit,
                "fundingAmount": funding_amount,
                "fundingIncomeTime": funding_ms,
                "openOffsetMs": open_offset,
                "closeOffsetMs": summary["closeTime"] - target_ms,
            })
            time.sleep(REQUEST_DELAY_SEC)

    results.sort(key=lambda row: int(row["openTime"]))
    write_excel(results, output_path)
    print(f"已输出: {output_path}")
    print(f"样本: {len(results)} 标的: {len(symbols)}")


if __name__ == "__main__":
    main(
        days=30,                       # 查询最近多少天
        open_before_hour_ms=3000,      # 只看开仓在 funding 前多少毫秒内
        before_sec=30,                 # funding 前多少秒计算平均波动率
        after_sec=3,                   # funding 后多少秒逐秒展示波动率
        funding_tolerance_ms=120000,   # funding 入账时间允许偏离 T 多少毫秒
        include_excluded=False,        # 是否包含 maxfunding 排除列表里的币
        output_path=ROOT / "tools" / "fundingresultcheck.xlsx", # Excel 输出路径
    )
