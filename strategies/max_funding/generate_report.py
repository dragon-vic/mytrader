from __future__ import annotations

import hashlib
import hmac
import os
import platform
import time
from datetime import datetime
from datetime import timezone
from decimal import Decimal
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font


STRATEGY_DIR = Path(__file__).resolve().parent
ROOT = STRATEGY_DIR.parents[1]
BASE_URL = "https://fapi.binance.com"
OUTPUT_PATH = STRATEGY_DIR / "maxfunding_event_report.xlsx"
EVENT_PATH = STRATEGY_DIR / "strategy_events.csv"
REQUEST_DELAY_SEC = 0.1
QUERY_BEFORE_MS = 10_000
QUERY_AFTER_MS = 30_000
FUNDING_TOLERANCE_MS = 10
TICK_BEFORE_MS = 3000
TICK_AFTER_MS = 3000
REQUEST_RETRIES = 3


def format_time(ms: int | None) -> str:
    if ms is None:
        return "-"
    dt = datetime.fromtimestamp(ms / 1000, tz=timezone.utc).astimezone()
    return dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]


def format_hour(ms: int) -> str:
    dt = datetime.fromtimestamp(ms / 1000, tz=timezone.utc).astimezone()
    return dt.strftime("%Y-%m-%d %H:00")


def offset_label(offset_ms: int | None) -> str:
    if offset_ms is None:
        return "-"
    sign = "+" if offset_ms >= 0 else "-"
    return f"T{sign}{abs(offset_ms)}ms"


def parse_offset(value: Any) -> int | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if not text or text == "-":
        return None
    if not text.startswith("T") or not text.endswith("ms"):
        return None
    sign = -1 if text[1:2] == "-" else 1
    digits = text[2:-2] if text[1:2] in {"+", "-"} else text[1:-2]
    try:
        return sign * int(digits)
    except ValueError:
        return None


def event_time_ms(value: Any) -> int:
    return int(pd.Timestamp(value).timestamp() * 1000)


def latest_event_path() -> Path | None:
    return EVENT_PATH if EVENT_PATH.exists() else None


def load_events() -> pd.DataFrame:
    path = latest_event_path()
    if path is None:
        return pd.DataFrame(columns=["symbol", "funding_time", "close_order_submit_time"])
    events = pd.read_csv(path)
    if events.empty:
        return events
    events["symbol"] = events["symbol"].astype(str).str.upper()
    events["funding_ms"] = events["funding_time"].map(event_time_ms).astype("int64")
    events["close_submit_offset_ms"] = events["close_order_submit_time"].map(parse_offset)
    events["close_submit_ms"] = events["funding_ms"] + events["close_submit_offset_ms"].fillna(0).astype("int64")
    return events


def build_session() -> tuple[requests.Session, bytes]:
    load_dotenv(ROOT / ".env")
    session = requests.Session()
    session.headers.update({"X-MBX-APIKEY": os.environ["BINANCE_FUTURES_API_KEY"]})
    if platform.system() == "Windows":
        proxy_url = os.environ.get("PROXY_URL")
        if proxy_url:
            session.proxies.update({"http": proxy_url, "https": proxy_url})
    return session, os.environ["BINANCE_FUTURES_API_SECRET"].encode()


def request_json(session: requests.Session, path: str, params: dict[str, Any] | None = None) -> Any:
    last_error: Exception | None = None
    for attempt in range(1, REQUEST_RETRIES + 1):
        try:
            response = session.get(f"{BASE_URL}{path}", params=params or {}, timeout=30)
            data = response.json()
            if response.status_code in (418, 429):
                raise RuntimeError(f"RATE_LIMIT HTTP {response.status_code}: {data}")
            if response.status_code >= 400:
                raise RuntimeError(f"HTTP {response.status_code}: {data}")
            return data
        except (requests.RequestException, ValueError, RuntimeError) as exc:
            last_error = exc
            if str(exc).startswith("RATE_LIMIT"):
                break
            if attempt >= REQUEST_RETRIES:
                break
            print(f"API请求失败，重试：{attempt}/{REQUEST_RETRIES}，路径：{path}，原因：{exc}", flush=True)
            time.sleep(0.5 * attempt)
    raise RuntimeError(f"API请求失败，路径：{path}，原因：{last_error}")


def signed_request(
    session: requests.Session,
    secret: bytes,
    path: str,
    params: dict[str, Any] | None = None,
) -> list[dict[str, Any]] | dict[str, Any]:
    last_error: Exception | None = None
    for attempt in range(1, REQUEST_RETRIES + 1):
        try:
            payload = dict(params or {})
            payload["timestamp"] = int(time.time() * 1000)
            payload["recvWindow"] = 10000
            query = urlencode(payload)
            sig = hmac.new(secret, query.encode(), hashlib.sha256).hexdigest()
            response = session.get(f"{BASE_URL}{path}?{query}&signature={sig}", timeout=30)
            data = response.json()
            if response.status_code in (418, 429):
                raise RuntimeError(f"RATE_LIMIT HTTP {response.status_code}: {data}")
            if response.status_code >= 400:
                raise RuntimeError(f"HTTP {response.status_code}: {data}")
            return data
        except (requests.RequestException, ValueError, RuntimeError) as exc:
            last_error = exc
            if str(exc).startswith("RATE_LIMIT"):
                break
            if attempt >= REQUEST_RETRIES:
                break
            print(f"API请求失败，重试：{attempt}/{REQUEST_RETRIES}，路径：{path}，原因：{exc}", flush=True)
            time.sleep(0.5 * attempt)
    raise RuntimeError(f"API请求失败，路径：{path}，原因：{last_error}")


def signed_pages(
    session: requests.Session,
    secret: bytes,
    path: str,
    params: dict[str, Any],
    start_ms: int,
    end_ms: int,
    key: str = "time",
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current = start_ms
    for _ in range(20):
        payload = dict(params)
        payload.update({"startTime": current, "endTime": end_ms, "limit": 1000})
        batch = signed_request(session, secret, path, payload)
        if not isinstance(batch, list) or not batch:
            break
        batch.sort(key=lambda item: int(item[key]))
        rows.extend(batch)
        last = int(batch[-1][key])
        if len(batch) < 1000 or last >= end_ms:
            break
        current = last + 1
        time.sleep(REQUEST_DELAY_SEC)

    seen: set[tuple[Any, ...]] = set()
    result = []
    for row in rows:
        ident = (row.get("symbol"), row.get("id"), row.get("orderId"), row.get("time"))
        if ident not in seen:
            seen.add(ident)
            result.append(row)
    return sorted(result, key=lambda item: int(item[key]))


def user_trades(session: requests.Session, secret: bytes, symbol: str, start_ms: int, end_ms: int) -> list[dict[str, Any]]:
    return signed_pages(session, secret, "/fapi/v1/userTrades", {"symbol": symbol}, start_ms, end_ms)


def api_funding_row(session: requests.Session, symbol: str, funding_ms: int) -> dict[str, Any]:
    data = request_json(
        session,
        "/fapi/v1/fundingRate",
        {
            "symbol": symbol,
            "startTime": funding_ms - FUNDING_TOLERANCE_MS,
            "endTime": funding_ms + FUNDING_TOLERANCE_MS,
            "limit": 10,
        },
    )
    if not isinstance(data, list) or not data:
        return {}
    row = min(data, key=lambda item: abs(int(item["fundingTime"]) - funding_ms))
    if abs(int(row["fundingTime"]) - funding_ms) > FUNDING_TOLERANCE_MS:
        return {}
    rate = float(row["fundingRate"])
    return {
        "symbol": symbol,
        "funding_time": int(row["fundingTime"]),
        "rate": rate,
        "rate_bps": rate * 10000.0,
        "side": "SELL" if rate > 0 else "BUY",
    }


def api_ticks(session: requests.Session, symbol: str, funding_ms: int) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    cursor = funding_ms - TICK_BEFORE_MS
    end_ms = funding_ms + TICK_AFTER_MS
    while cursor <= end_ms:
        data = request_json(
            session,
            "/fapi/v1/aggTrades",
            {"symbol": symbol, "startTime": cursor, "endTime": end_ms, "limit": 1000},
        )
        if not isinstance(data, list) or not data:
            break
        rows.extend(data)
        last_ts = max(int(item["T"]) for item in data)
        if len(data) < 1000 or last_ts >= end_ms:
            break
        cursor = last_ts + 1
        time.sleep(REQUEST_DELAY_SEC)
    if not rows:
        return pd.DataFrame(columns=["symbol", "timestamp_ms", "price"])
    frame = pd.DataFrame(rows).rename(columns={"T": "timestamp_ms", "p": "price"})
    frame["symbol"] = symbol
    frame["timestamp_ms"] = pd.to_numeric(frame["timestamp_ms"], errors="coerce").astype("int64")
    frame["price"] = pd.to_numeric(frame["price"], errors="coerce")
    return frame[["symbol", "timestamp_ms", "price"]].dropna().drop_duplicates()


def trade_delta(trade: dict[str, Any]) -> Decimal:
    qty = Decimal(str(trade["qty"]))
    side = trade["side"]
    position_side = trade.get("positionSide", "BOTH")
    if position_side == "SHORT":
        return qty if side == "SELL" else -qty
    return qty if side == "BUY" else -qty


def closed_segments(trades: list[dict[str, Any]], start_ms: int, end_ms: int) -> list[dict[str, Any]]:
    segments = []
    sides = sorted({trade.get("positionSide", "BOTH") for trade in trades})
    for position_side in sides:
        pos = Decimal("0")
        current: dict[str, Any] | None = None
        side_trades = [trade for trade in trades if trade.get("positionSide", "BOTH") == position_side]
        for trade in side_trades:
            before = pos
            after = before + trade_delta(trade)
            if before == 0 and after != 0:
                current = {
                    "symbol": trade["symbol"],
                    "positionSide": position_side,
                    "direction": "LONG" if after > 0 else "SHORT",
                    "openTime": int(trade["time"]),
                    "trades": [],
                }
            if current is not None:
                current["trades"].append(trade)
            flipped = before != 0 and ((before > 0 and after < 0) or (before < 0 and after > 0))
            if before != 0 and (after == 0 or flipped) and current is not None:
                current["closeTime"] = int(trade["time"])
                segments.append(current)
                current = (
                    {
                        "symbol": trade["symbol"],
                        "positionSide": position_side,
                        "direction": "LONG" if after > 0 else "SHORT",
                        "openTime": int(trade["time"]),
                        "trades": [trade],
                    }
                    if after != 0
                    else None
                )
            pos = after
    return sorted(
        [segment for segment in segments if start_ms <= int(segment["closeTime"]) <= end_ms],
        key=lambda item: int(item["closeTime"]),
    )


def summarize_segment(segment: dict[str, Any]) -> dict[str, Any]:
    pos = Decimal("0")
    max_notional = Decimal("0")
    pnl = Decimal("0")
    for trade in segment["trades"]:
        pos += trade_delta(trade)
        max_notional = max(max_notional, abs(pos) * Decimal(str(trade["price"])))
        pnl += Decimal(str(trade.get("realizedPnl", "0")))
    return {**segment, "notional": max_notional, "realizedPnl": pnl}


def match_segment(segments: list[dict[str, Any]], funding_ms: int, close_submit_ms: int | None) -> tuple[dict[str, Any] | None, str]:
    covering = [segment for segment in segments if int(segment["openTime"]) <= funding_ms <= int(segment["closeTime"])]
    if covering:
        return covering[0], "匹配"
    if close_submit_ms is None or not segments:
        return None, "未匹配"
    return min(segments, key=lambda segment: abs(int(segment["closeTime"]) - close_submit_ms)), "弱匹配"


def nearest_price(ticks: pd.DataFrame, target_ms: int) -> float | None:
    if ticks.empty:
        return None
    ticks = ticks.copy()
    ticks["delta"] = (ticks["timestamp_ms"].astype("int64") - target_ms).abs()
    row = ticks.sort_values("delta").iloc[0]
    return float(row["price"])


def tick_stats(ticks: pd.DataFrame, symbol: str, funding_ms: int, side: str) -> dict[str, Any]:
    if ticks.empty:
        return {"tick_count": 0, "pre_price": None, "post_price": None, "slip_bps": None}
    window = ticks[
        ticks["symbol"].astype(str).str.upper().eq(symbol)
        & ticks["timestamp_ms"].astype("int64").between(funding_ms - 3000, funding_ms + 3000)
    ].copy()
    pre = nearest_price(window, funding_ms - 3000)
    post = nearest_price(window, funding_ms + 2000)
    slip = None
    if pre is not None and post is not None and pre:
        direction = -1.0 if side == "SELL" else 1.0
        slip = direction * (post - pre) / pre * 10000.0
    return {"tick_count": len(window), "pre_price": pre, "post_price": post, "slip_bps": slip}


def report_rows(events: pd.DataFrame) -> list[dict[str, Any]]:
    if events.empty:
        return []

    session, secret = build_session()
    rows = []
    total = len(events)
    for index, event in enumerate(events.itertuples(index=False), start=1):
        symbol = str(event.symbol).upper()
        funding_ms = int(event.funding_ms)
        close_submit_ms = int(event.close_submit_ms) if pd.notna(event.close_submit_ms) else None
        print(f"处理事件：{index}/{total}，交易对：{symbol}，T时间：{format_hour(funding_ms)}", flush=True)
        frow = api_funding_row(session, symbol, funding_ms)
        side = str(frow.get("side") or "")
        if side not in {"BUY", "SELL"}:
            rate_bps = float(frow.get("rate_bps", 0.0) or 0.0)
            side = "SELL" if rate_bps > 0 else "BUY"

        ticks = api_ticks(session, symbol, funding_ms)
        trades = user_trades(session, secret, symbol, funding_ms - QUERY_BEFORE_MS, funding_ms + QUERY_AFTER_MS)
        segments = closed_segments(trades, funding_ms - QUERY_BEFORE_MS, funding_ms + QUERY_AFTER_MS)
        segment, status = match_segment(segments, funding_ms, close_submit_ms)
        summary = summarize_segment(segment) if segment is not None else {}
        tick = tick_stats(ticks, symbol, funding_ms, side)
        api_open = int(summary["openTime"]) if summary else None
        api_close = int(summary["closeTime"]) if summary else None
        rows.append(
            {
                "funding_ms": funding_ms,
                "symbol": symbol,
                "rate_bps": frow.get("rate_bps"),
                "side": side,
                "event_close_offset": int(event.close_submit_offset_ms) if pd.notna(event.close_submit_offset_ms) else None,
                "api_open_offset": api_open - funding_ms if api_open is not None else None,
                "api_close_offset": api_close - funding_ms if api_close is not None else None,
                "submit_fill_delay": api_close - close_submit_ms if api_close is not None and close_submit_ms is not None else None,
                "notional": summary.get("notional"),
                "realized_pnl": summary.get("realizedPnl"),
                "tick_count": tick["tick_count"],
                "pre_price": tick["pre_price"],
                "post_price": tick["post_price"],
                "slip_bps": tick["slip_bps"],
                "status": status,
            },
        )
        time.sleep(REQUEST_DELAY_SEC)
    return rows


def write_excel(rows: list[dict[str, Any]]) -> None:
    headers = [
        "T时间",
        "标的",
        "funding_rate_bps",
        "方向",
        "event平仓提交偏移",
        "API开仓偏移",
        "API平仓偏移",
        "提交到成交延迟ms",
        "名义价值",
        "realized_pnl",
        "funding附近tick数",
        "T前3秒价格",
        "T后2秒价格",
        "T后2秒滑点bps",
        "匹配状态",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "maxfunding结果"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        notional = row.get("notional")
        realized = row.get("realized_pnl")
        ws.append(
            [
                format_hour(int(row["funding_ms"])),
                row["symbol"],
                row.get("rate_bps"),
                row.get("side"),
                offset_label(row.get("event_close_offset")),
                offset_label(row.get("api_open_offset")),
                offset_label(row.get("api_close_offset")),
                row.get("submit_fill_delay"),
                float(notional) if isinstance(notional, Decimal) else notional,
                float(realized) if isinstance(realized, Decimal) else realized,
                row.get("tick_count"),
                row.get("pre_price"),
                row.get("post_price"),
                row.get("slip_bps"),
                row.get("status"),
            ],
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
    widths = {
        "A": 18,
        "B": 14,
        "C": 18,
        "D": 8,
        "E": 18,
        "F": 14,
        "G": 14,
        "H": 18,
        "I": 12,
        "J": 12,
        "K": 16,
        "L": 12,
        "M": 12,
        "N": 16,
        "O": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)


def main() -> Path:
    print("读取策略 events", flush=True)
    events = load_events()
    print(f"开始生成 Excel，events：{len(events)} 行", flush=True)
    rows = report_rows(events)
    print(f"API 查询完成，报表行数：{len(rows)}", flush=True)
    write_excel(rows)
    print(f"Excel输出完成，行数：{len(rows)}，文件：{OUTPUT_PATH}", flush=True)
    return OUTPUT_PATH


if __name__ == "__main__":
    main()
